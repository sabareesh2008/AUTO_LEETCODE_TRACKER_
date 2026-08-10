from __future__ import annotations

import csv
import os
from datetime import date, datetime, timedelta, time as dt_time
from zoneinfo import ZoneInfo
from pathlib import Path
from typing import Any

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ============================================================
# SETTINGS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent

STUDENTS_FILE = BASE_DIR / "students.csv"
LIVE_CSV = BASE_DIR / "LiveData.csv"
HISTORY_CSV = BASE_DIR / "History.csv"
DAILY_ACTIVITY_CSV = BASE_DIR / "DailyActivity.csv"
STUDENTS_XLSX = BASE_DIR / "Students.xlsx"

SUPABASE_URL = os.getenv("SUPABASE_URL", "").rstrip("/")
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")

LEETCODE_URL = "https://leetcode.com/graphql"

# Large enough for most normal student activity.
# If a student has more than this many accepted submissions
# inside the requested period, the 7/30 day counts can undercount.
RECENT_SUBMISSION_LIMIT = 300
IST = ZoneInfo("Asia/Kolkata")


LEETCODE_QUERY = """
query getUserProfile($username: String!, $limit: Int!) {
  matchedUser(username: $username) {
    username
    submitStatsGlobal {
      acSubmissionNum {
        difficulty
        count
        submissions
      }
    }
  }

  recentAcSubmissionList(
    username: $username,
    limit: $limit
  ) {
    title
    titleSlug
    timestamp
  }
}
"""


# ============================================================
# HELPERS
# ============================================================

def clean(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def empty_profile(status: str) -> dict[str, Any]:
    return {
        "total_solved": 0,
        "easy": 0,
        "medium": 0,
        "hard": 0,
        "submissions": 0,
        "solved_today": 0,
        "last_7_days": 0,
        "last_30_days": 0,
        "last_problem": "",
        "last_solved": "",
        "status": status,
        "recent_submissions": [],
    }


def get_stat(
    statistics: list[dict[str, Any]],
    difficulty: str,
    field: str = "count",
) -> int:
    for statistic in statistics:
        if statistic.get("difficulty") == difficulty:
            return int(statistic.get(field, 0) or 0)
    return 0


def unique_solved_count_since(
    recent_submissions: list[dict[str, Any]],
    start_time: datetime,
) -> int:
    solved = set()

    for submission in recent_submissions:
        timestamp = submission.get("timestamp")
        title_slug = submission.get("titleSlug")

        if not timestamp or not title_slug:
            continue

        submission_time = datetime.fromtimestamp(int(timestamp))

        if submission_time >= start_time:
            solved.add(title_slug)

    return len(solved)


# ============================================================
# LEETCODE FETCH
# ============================================================

def fetch_leetcode(username: str) -> dict[str, Any]:
    if not username:
        return empty_profile("Username missing")

    request_body = {
        "operationName": "getUserProfile",
        "query": LEETCODE_QUERY,
        "variables": {
            "username": username,
            "limit": RECENT_SUBMISSION_LIMIT,
        },
    }

    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "Origin": "https://leetcode.com",
        "Referer": f"https://leetcode.com/u/{username}/",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/151.0.0.0 Safari/537.36"
        ),
    }

    try:
        response = requests.post(
            LEETCODE_URL,
            json=request_body,
            headers=headers,
            timeout=30,
        )

        if response.status_code != 200:
            return empty_profile(f"HTTP {response.status_code}")

        try:
            response_data = response.json()
        except ValueError:
            return empty_profile("Invalid JSON response")

        if response_data.get("errors"):
            messages = [
                str(error.get("message", "GraphQL error"))
                for error in response_data["errors"]
            ]
            return empty_profile(" | ".join(messages))

        data = response_data.get("data", {})
        matched_user = data.get("matchedUser")

        if matched_user is None:
            return empty_profile("User not found")

        statistics = (
            matched_user
            .get("submitStatsGlobal", {})
            .get("acSubmissionNum", [])
        )

        recent_submissions = data.get("recentAcSubmissionList", []) or []

        last_problem = ""
        last_solved = ""

        if recent_submissions:
            latest = recent_submissions[0]
            last_problem = clean(latest.get("title"))

            timestamp = latest.get("timestamp")
            if timestamp:
                last_solved = datetime.fromtimestamp(
                    int(timestamp)
                ).strftime("%Y-%m-%d %H:%M:%S")

        now = datetime.now()

        today_start = datetime.combine(date.today(), datetime.min.time())
        seven_days_start = now - timedelta(days=7)
        thirty_days_start = now - timedelta(days=30)

        return {
            "total_solved": get_stat(statistics, "All"),
            "easy": get_stat(statistics, "Easy"),
            "medium": get_stat(statistics, "Medium"),
            "hard": get_stat(statistics, "Hard"),
            "submissions": get_stat(statistics, "All", "submissions"),
            "solved_today": unique_solved_count_since(
                recent_submissions,
                today_start,
            ),
            "last_7_days": unique_solved_count_since(
                recent_submissions,
                seven_days_start,
            ),
            "last_30_days": unique_solved_count_since(
                recent_submissions,
                thirty_days_start,
            ),
            "last_problem": last_problem,
            "last_solved": last_solved,
            "status": "Success",
            "recent_submissions": recent_submissions,
        }

    except requests.Timeout:
        return empty_profile("Request timeout")

    except requests.ConnectionError as error:
        return empty_profile(f"Connection error: {error}")

    except requests.RequestException as error:
        return empty_profile(f"Network error: {error}")

    except Exception as error:
        return empty_profile(f"Unexpected error: {error}")


# ============================================================
# INPUT
# ============================================================

def sync_students_from_supabase() -> None:
    """Download the authoritative student directory from Supabase.

    On GitHub Actions the two secrets are required. For local/offline testing,
    if they are missing, the existing students.csv is used instead.
    """
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        print("Supabase secrets not set; using local students.csv")
        return

    url = f"{SUPABASE_URL}/rest/v1/students"
    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "Accept": "application/json",
    }
    params = {
        "select": "register_number,student_name,leetcode_username,section,created_at",
        "order": "created_at.asc",
    }
    response = requests.get(url, headers=headers, params=params, timeout=30)
    response.raise_for_status()
    rows = response.json()

    frame = pd.DataFrame([{
        "Register Number": clean(row.get("register_number")),
        "Student Name": clean(row.get("student_name")),
        "LeetCode Username": clean(row.get("leetcode_username")),
        "Section": clean(row.get("section")),
    } for row in rows])

    if frame.empty:
        frame = pd.DataFrame(columns=["Register Number", "Student Name", "LeetCode Username", "Section"])

    atomic_csv_write(frame, STUDENTS_FILE)
    print(f"Synced {len(frame)} student(s) from Supabase")


def write_students_excel(students: pd.DataFrame) -> None:
    """Create a clean Excel copy automatically; users never edit it manually."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    headers = ["Register Number", "Student Name", "LeetCode Username", "LeetCode Link"]
    ws.append(headers)

    for _, student in students.iterrows():
        username = clean(student["LeetCode Username"])
        ws.append([
            clean(student["Register Number"]),
            clean(student["Student Name"]),
            username,
            f"https://leetcode.com/u/{username}/",
        ])

    header_fill = PatternFill("solid", fgColor="2563EB")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    widths = {"A": 20, "B": 28, "C": 28, "D": 48}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(STUDENTS_XLSX)


def read_students() -> pd.DataFrame:
    sync_students_from_supabase()

    if not STUDENTS_FILE.exists():
        raise FileNotFoundError(f"students.csv not found: {STUDENTS_FILE}")

    students = pd.read_csv(STUDENTS_FILE, dtype=str, keep_default_na=False)
    required_columns = {"Register Number", "Student Name", "LeetCode Username"}
    if "Section" not in students.columns:
        students["Section"] = ""
    missing = required_columns.difference(students.columns)
    if missing:
        raise ValueError("Missing students.csv columns: " + ", ".join(sorted(missing)))

    students = students.dropna(how="all").copy()
    for column in required_columns:
        students[column] = students[column].apply(clean)
    students = students[(students["Student Name"] != "") & (students["LeetCode Username"] != "")].copy()
    write_students_excel(students)
    return students


# ============================================================
# HISTORY
# ============================================================

HISTORY_COLUMNS = [
    "Date",
    "Register Number",
    "Student Name",
    "LeetCode Username",
    "Problems Solved",
    "Solved Today",
    "Last 7 Days",
    "Last 30 Days",
    "Total Submissions",
    "Easy",
    "Medium",
    "Hard",
    "Last Problem",
    "Last Solved",
    "Status",
    "Updated At",
]


def load_history() -> pd.DataFrame:
    if not HISTORY_CSV.exists():
        return pd.DataFrame(columns=HISTORY_COLUMNS)

    try:
        history = pd.read_csv(
            HISTORY_CSV,
            dtype=str,
            keep_default_na=False,
        )
    except pd.errors.EmptyDataError:
        return pd.DataFrame(columns=HISTORY_COLUMNS)

    for column in HISTORY_COLUMNS:
        if column not in history.columns:
            history[column] = ""

    return history[HISTORY_COLUMNS].copy()


def update_history(
    previous_history: pd.DataFrame,
    current_rows: list[dict[str, Any]],
) -> pd.DataFrame:
    today_text = date.today().isoformat()

    history = previous_history.copy()

    if not history.empty:
        # Replace today's snapshot for each current student so frequent runs
        # do not create duplicate rows.
        current_registers = {
            str(row["Register Number"])
            for row in current_rows
        }

        history = history[
            ~(
                (history["Date"] == today_text)
                & history["Register Number"].astype(str).isin(current_registers)
            )
        ].copy()

    new_history_rows = []

    for row in current_rows:
        new_history_rows.append({
            "Date": today_text,
            "Register Number": row["Register Number"],
            "Student Name": row["Student Name"],
            "LeetCode Username": row["LeetCode Username"],
            "Problems Solved": row["Problems Solved"],
            "Solved Today": row["Solved Today"],
            "Last 7 Days": row["Last 7 Days"],
            "Last 30 Days": row["Last 30 Days"],
            "Total Submissions": row["Total Submissions"],
            "Easy": row["Easy"],
            "Medium": row["Medium"],
            "Hard": row["Hard"],
            "Last Problem": row["Last Problem"],
            "Last Solved": row["Last Solved"],
            "Status": row["Status"],
            "Updated At": row["Updated At"],
        })

    combined = pd.concat(
        [
            history,
            pd.DataFrame(new_history_rows, columns=HISTORY_COLUMNS),
        ],
        ignore_index=True,
    )

    if not combined.empty:
        combined = combined.sort_values(
            by=["Date", "Student Name"],
            ascending=[True, True],
        ).reset_index(drop=True)

    return combined


def build_daily_activity(history: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "Date",
        "Register Number",
        "Student Name",
        "LeetCode Username",
        "Problems Solved",
        "Solved That Day",
    ]

    if history.empty:
        return pd.DataFrame(columns=columns)

    activity = history[
        [
            "Date",
            "Register Number",
            "Student Name",
            "LeetCode Username",
            "Problems Solved",
            "Solved Today",
        ]
    ].copy()

    activity = activity.rename(
        columns={"Solved Today": "Solved That Day"}
    )

    return activity[columns]


# ============================================================
# SAFE CSV WRITE
# ============================================================

def atomic_csv_write(
    dataframe: pd.DataFrame,
    destination: Path,
) -> None:
    temporary_file = destination.with_suffix(".temporary.csv")

    dataframe.to_csv(
        temporary_file,
        index=False,
        encoding="utf-8-sig",
    )

    os.replace(
        temporary_file,
        destination,
    )



# ============================================================
# DAILY CHALLENGE
# ============================================================

def supabase_headers() -> dict[str, str]:
    return {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }


def load_recent_challenges(days: int = 35) -> list[dict[str, Any]]:
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        return []

    start_date = (datetime.now(IST).date() - timedelta(days=days)).isoformat()
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/daily_challenges",
        headers=supabase_headers(),
        params={
            "select": "id,challenge_date,problem_title,problem_slug,problem_url,difficulty",
            "challenge_date": f"gte.{start_date}",
            "order": "challenge_date.asc",
        },
        timeout=30,
    )
    response.raise_for_status()
    return response.json()


def challenge_completion(
    recent_submissions: list[dict[str, Any]],
    challenge: dict[str, Any],
) -> tuple[bool, str | None]:
    challenge_day = date.fromisoformat(str(challenge["challenge_date"]))
    target_slug = clean(challenge.get("problem_slug")).lower()

    for submission in recent_submissions:
        if clean(submission.get("titleSlug")).lower() != target_slug:
            continue

        timestamp = submission.get("timestamp")
        if not timestamp:
            continue

        solved_at = datetime.fromtimestamp(int(timestamp), tz=IST)

        # A daily challenge counts only when solved on the assigned IST date.
        if solved_at.date() == challenge_day:
            return True, solved_at.isoformat()

    return False, None


def save_challenge_result(
    challenge_id: int,
    register_number: str,
    completed: bool,
    completed_at: str | None,
) -> None:
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        return

    payload = {
        "challenge_id": challenge_id,
        "register_number": register_number,
        "completed": completed,
        "completed_at": completed_at,
        "checked_at": datetime.now(IST).isoformat(),
    }

    response = requests.post(
        f"{SUPABASE_URL}/rest/v1/daily_challenge_results",
        headers={
            **supabase_headers(),
            "Prefer": "resolution=merge-duplicates,return=minimal",
        },
        params={"on_conflict": "challenge_id,register_number"},
        json=payload,
        timeout=30,
    )
    response.raise_for_status()


# ============================================================
# MAIN UPDATE
# ============================================================

def run_one_update() -> None:
    students = read_students()

    updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    live_rows: list[dict[str, Any]] = []

    total_students = len(students)
    recent_challenges = load_recent_challenges()

    print("=" * 64)
    print(f"LeetCode cloud update: {updated_at}")
    print(f"Students: {total_students}")
    print("=" * 64)

    for position, (_, student) in enumerate(
        students.iterrows(),
        start=1,
    ):
        register_number = clean(student["Register Number"])
        student_name = clean(student["Student Name"])
        username = clean(student["LeetCode Username"])

        print(
            f"[{position}/{total_students}] "
            f"{student_name} ({username})"
        )

        profile = fetch_leetcode(username)

        for challenge in recent_challenges:
            completed, completed_at = challenge_completion(
                profile.get("recent_submissions", []),
                challenge,
            )
            save_challenge_result(
                int(challenge["id"]),
                register_number,
                completed,
                completed_at,
            )

        live_rows.append({
            "Register Number": register_number,
            "Student Name": student_name,
            "Section": clean(student.get("Section", "")),
            "LeetCode Username": username,
            "LeetCode Link": (
                f"https://leetcode.com/u/{username}/"
            ),
            "Problems Solved": profile["total_solved"],
            "Solved Today": profile["solved_today"],
            "Last 7 Days": profile["last_7_days"],
            "Last 30 Days": profile["last_30_days"],
            "Total Submissions": profile["submissions"],
            "Easy": profile["easy"],
            "Medium": profile["medium"],
            "Hard": profile["hard"],
            "Last Problem": profile["last_problem"],
            "Last Solved": profile["last_solved"],
            "Status": profile["status"],
            "Updated At": updated_at,
        })

        print(
            "  "
            f"30d={profile['last_30_days']} | "
            f"7d={profile['last_7_days']} | "
            f"today={profile['solved_today']} | "
            f"total={profile['total_solved']} | "
            f"{profile['status']}"
        )

    live_data = pd.DataFrame(live_rows)

    if not live_data.empty:
        # Rank is based on last 30 days.
        # Tie-breakers: last 7 days, solved today, total solved, name.
        live_data = live_data.sort_values(
            by=[
                "Last 30 Days",
                "Last 7 Days",
                "Solved Today",
                "Problems Solved",
                "Student Name",
            ],
            ascending=[
                False,
                False,
                False,
                False,
                True,
            ],
        ).reset_index(drop=True)

        live_data.insert(
            0,
            "Rank",
            range(1, len(live_data) + 1),
        )

    history = update_history(
        load_history(),
        live_rows,
    )

    daily_activity = build_daily_activity(history)

    atomic_csv_write(
        live_data,
        LIVE_CSV,
    )

    atomic_csv_write(
        history,
        HISTORY_CSV,
    )

    atomic_csv_write(
        daily_activity,
        DAILY_ACTIVITY_CSV,
    )

    print("=" * 64)
    print("CSV files updated successfully.")
    print(f"LiveData.csv: {LIVE_CSV}")
    print(f"History.csv: {HISTORY_CSV}")
    print(f"DailyActivity.csv: {DAILY_ACTIVITY_CSV}")
    print("=" * 64)


if __name__ == "__main__":
    run_one_update()
